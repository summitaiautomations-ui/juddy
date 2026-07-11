#!/usr/bin/env python3
"""Build 'Disc Diver Sales Tracker.xlsx' from inventory.csv + orders.csv
   (+ optional customers.csv).

Three tabs:
  * "Orders"     -- every sold disc, with Buyer / Paid / Paid Via / Shipped /
                    Tracking. Paid? & Shipped? are Yes/No dropdowns that turn
                    green when done. A summary up top totals your money.
  * "Customers"  -- your little CRM: one row per buyer with Cell, Shipping
                    Address, Pays With, and what molds/brands they Like, plus
                    auto-calculated # Orders / Total Spent / Repeat?. Fill in
                    the contact fields so you can text past buyers when you
                    find a disc they'd want -> repeat sales.
  * "For Sale"   -- everything still available, so you can see what's left.

Re-run this any time the inventory, orders, or customers change. Safe to run
repeatedly. orders.csv and customers.csv are PRIVATE (gitignored) -- they hold
buyer names / cells / addresses and never get pushed to the public repo.
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA = os.path.join(ROOT, "disc-pics-data")
OUT = os.path.join(DATA, "Disc Diver Sales Tracker.xlsx")

TEAL = "0E5A66"; TEAL_D = "093C45"; PLUM = "5B3A66"
GREEN = "D6F0DC"; GREEN_T = "1E7A34"; HEADFG = "FFFFFF"
GREY = "6B7A80"
thin = Side(style="thin", color="D9DEE1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def rows_csv(name):
    path = os.path.join(DATA, name)
    if not os.path.exists(path):
        return []
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def disc_name(d):
    parts = [d["brand"],
             d["plastic"] if d["plastic"] not in ("", "unknown") else "",
             d["mold"] if d["mold"] not in ("", "unknown") else "Disc"]
    return " ".join(p for p in parts if p)


def weight(d):
    for k in ("scale_weight", "stamped_weight"):
        v = d.get(k, "")
        if v and v.replace(".", "").isdigit():
            return f"{v}g"
    return ""


def price_num(d):
    try:
        return float(d["price"])
    except (ValueError, KeyError):
        return 0.0


def style_header(ws, headers, fill, row):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color=HEADFG, size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


GREEN_FILL = PatternFill("solid", fgColor=GREEN)
GREEN_FONT = Font(color=GREEN_T, bold=True)


def yesno_green(ws, col, r0, r1):
    """Dropdown + green-when-Yes for one column between rows r0..r1."""
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"{col}{r0}:{col}{r1}")
    ws.conditional_formatting.add(
        f"{col}{r0}:{col}{r1}",
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=GREEN_FILL, font=GREEN_FONT))


def main():
    inv = {d["id"]: d for d in rows_csv("inventory.csv")}
    orders = {o["id"]: o for o in rows_csv("orders.csv")}
    customers = {c["name"]: c for c in rows_csv("customers.csv") if c.get("name")}

    wb = Workbook()

    # ==================== Orders tab ====================
    ws = wb.active
    ws.title = "Orders"
    ws.sheet_view.showGridLines = False

    ws.cell(1, 1, "DISC DIVER — SALES TRACKER").font = Font(
        bold=True, size=16, color=TEAL_D)
    summ_row = 2

    headers = ["#", "Disc", "Color", "Weight", "Price", "Buyer",
               "Paid?", "Paid Via", "Shipped?", "Tracking #", "Sold Date", "Notes"]
    head_row = 4
    style_header(ws, headers, TEAL, head_row)
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    sold_ids = [i for i, d in inv.items() if d["status"] in ("sold", "claimed")]
    for i in orders:
        if i not in sold_ids:
            sold_ids.append(i)
    sold_ids.sort()

    # per-buyer aggregation for the Customers tab
    agg = {}  # name -> {"count", "spent", "pay": set, "discs": []}

    total = collected = to_ship = 0.0
    r = head_row
    for i in sold_ids:
        r += 1
        d = inv.get(i, {})
        o = orders.get(i, {})
        amt = (o.get("amount", "") or "").strip()
        p = float(amt) if amt else price_num(d)
        name = (o.get("desc", "") or "").strip() or (
            disc_name(d) if d else "(disc " + i + ")")
        buyer = (o.get("buyer", "") or "").strip()
        pay = (o.get("paid_via", "") or "").strip()
        total += p
        paid = (o.get("paid", "no") or "no").lower() == "yes"
        shipped = (o.get("shipped", "no") or "no").lower() == "yes"
        if paid:
            collected += p
        if not shipped:
            to_ship += 1
        if buyer and buyer.lower() != "cash sale (stranger)":
            a = agg.setdefault(buyer, {"count": 0, "spent": 0.0,
                                       "pay": set(), "discs": []})
            a["count"] += 1
            a["spent"] += p
            if pay:
                a["pay"].add(pay)
            a["discs"].append(name)
        vals = ["#" + i, name, d.get("color", ""), weight(d), p, buyer,
                "Yes" if paid else "No", pay,
                "Yes" if shipped else "No",
                o.get("tracking", ""), o.get("sold_date", ""),
                o.get("order_notes", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if c in (1, 4, 5, 7, 9) else "left",
                vertical="center")
        ws.cell(r, 5).number_format = '"$"#,##0'
    last = r

    yesno_green(ws, "G", head_row + 1, last)   # Paid?
    yesno_green(ws, "I", head_row + 1, last)    # Shipped?

    ws.cell(summ_row, 1, f"{len(sold_ids)} sold  •  ${total:,.0f} total  •  "
            f"${collected:,.0f} collected  •  {int(to_ship)} still to ship"
            ).font = Font(bold=True, size=12, color=GREY)

    for c, w in enumerate([7, 30, 14, 9, 8, 18, 9, 13, 10, 20, 12, 30], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ==================== Customers tab ====================
    wc = wb.create_sheet("Customers")
    wc.sheet_view.showGridLines = False
    wc.cell(1, 1, "CUSTOMERS — text them when you find a disc they'd want"
            ).font = Font(bold=True, size=16, color=PLUM)

    # union of everyone who has bought + anyone pre-seeded in customers.csv
    names = sorted(set(agg) | set(customers))
    wc.cell(2, 1, f"{len(names)} buyers  •  "
            f"${sum(a['spent'] for a in agg.values()):,.0f} lifetime sales"
            ).font = Font(bold=True, size=12, color=GREY)

    ch = ["Customer", "Cell", "Shipping Address", "Pays With",
          "Likes (molds / brands)", "# Orders", "Total Spent",
          "Repeat?", "Notes"]
    chr_ = 4
    style_header(wc, ch, PLUM, chr_)
    wc.freeze_panes = wc.cell(row=chr_ + 1, column=1)

    r = chr_
    for nm in names:
        r += 1
        a = agg.get(nm, {"count": 0, "spent": 0.0, "pay": set(), "discs": []})
        cust = customers.get(nm, {})
        pays = cust.get("pays_with", "") or " / ".join(sorted(a["pay"]))
        repeat = "Yes" if a["count"] >= 2 else "No"
        vals = [nm, cust.get("cell", ""), cust.get("address", ""), pays,
                cust.get("likes", ""), a["count"], a["spent"],
                repeat, cust.get("cust_notes", "")]
        for c, v in enumerate(vals, 1):
            cell = wc.cell(r, c, v)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if c in (6, 7, 8) else "left",
                vertical="center")
        wc.cell(r, 7).number_format = '"$"#,##0'
    clast = r
    if clast >= chr_ + 1:
        wc.conditional_formatting.add(
            f"H{chr_+1}:H{clast}",
            CellIsRule(operator="equal", formula=['"Yes"'],
                       fill=GREEN_FILL, font=GREEN_FONT))
    for c, w in enumerate([18, 16, 34, 16, 26, 9, 12, 9, 30], 1):
        wc.column_dimensions[get_column_letter(c)].width = w

    # ==================== For Sale tab ====================
    ws2 = wb.create_sheet("For Sale")
    ws2.sheet_view.showGridLines = False
    ws2.cell(1, 1, "STILL AVAILABLE").font = Font(bold=True, size=16, color=TEAL_D)
    avail = sorted([d for d in inv.values() if d["status"] == "available"],
                   key=lambda d: d["id"])
    ws2.cell(2, 1, f"{len(avail)} discs available  •  "
             f"${sum(price_num(d) for d in avail):,.0f} listed").font = Font(
        bold=True, size=12, color=GREY)
    h2 = ["#", "Disc", "Color", "Weight", "Price", "Notes"]
    hr2 = 4
    style_header(ws2, h2, TEAL_D, hr2)
    ws2.freeze_panes = ws2.cell(row=hr2 + 1, column=1)
    r = hr2
    for d in avail:
        r += 1
        vals = ["#" + d["id"], disc_name(d), d["color"], weight(d),
                price_num(d), d.get("notes", "")]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(r, c, v)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if c in (1, 4, 5) else "left",
                vertical="center")
        ws2.cell(r, 5).number_format = '"$"#,##0'
    for c, w in enumerate([7, 30, 16, 9, 8, 50], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  {len(sold_ids)} orders, ${total:,.0f} total, "
          f"${collected:,.0f} collected, {int(to_ship)} to ship, "
          f"{len(names)} customers")


if __name__ == "__main__":
    main()
